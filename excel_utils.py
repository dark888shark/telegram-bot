import openpyxl
from openpyxl.utils import get_column_letter

def append_to_excel(path, row_data):
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Имя", "Телефон", "Адрес", "Дата"])

    ws.append(row_data)

    for i, _ in enumerate(row_data, 1):
        ws.column_dimensions[get_column_letter(i)].width = 25

    wb.save(path)
